"""
setup_menu.py
Creates the data/menu.xlsx file with sample restaurant menu data.
Run this script once to initialize the menu spreadsheet.
"""

import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_menu_excel(output_path: str = "data/menu.xlsx") -> None:
    """Create the menu Excel file with sample data."""

    # Ensure the data directory exists
    data_dir = Path(output_path).parent
    data_dir.mkdir(parents=True, exist_ok=True)

    # Menu data: ID, Category, Name, Description, Price, Available, Dietary, Calories
    menu_items = [
        # ---------- Appetizers ----------
        ("A001", "Appetizers", "Crispy Calamari",
         "Tender calamari rings lightly breaded and fried to golden perfection, served with marinara and lemon aioli",
         12.99, True, "Seafood", 320),
        ("A002", "Appetizers", "Bruschetta al Pomodoro",
         "Grilled sourdough topped with fresh Roma tomatoes, garlic, basil, and a drizzle of extra-virgin olive oil",
         9.99, True, "Vegetarian, Vegan", 210),
        ("A003", "Appetizers", "Spinach & Artichoke Dip",
         "Creamy blend of spinach, artichoke hearts, and three cheeses baked until bubbly, served with tortilla chips",
         11.49, True, "Vegetarian, Gluten-Free", 380),
        ("A004", "Appetizers", "Buffalo Chicken Wings",
         "Crispy chicken wings tossed in classic buffalo sauce, served with blue cheese dip and celery sticks",
         14.99, True, "None", 490),
        ("A005", "Appetizers", "Shrimp Cocktail",
         "Chilled jumbo shrimp arranged with zesty cocktail sauce and fresh lemon wedges",
         16.99, True, "Seafood, Gluten-Free", 140),
        ("A006", "Appetizers", "Loaded Potato Skins",
         "Crispy potato skins filled with cheddar cheese, bacon bits, and green onions, served with sour cream",
         10.99, True, "None", 430),

        # ---------- Salads ----------
        ("S001", "Salads", "Classic Caesar Salad",
         "Crisp romaine lettuce tossed with housemade Caesar dressing, Parmesan shavings, and garlic croutons",
         13.49, True, "Vegetarian", 290),
        ("S002", "Salads", "Garden Fresh Salad",
         "Mixed greens, cherry tomatoes, cucumber, red onion, and carrot ribbons with your choice of dressing",
         10.99, True, "Vegetarian, Vegan, Gluten-Free", 180),
        ("S003", "Salads", "Greek Salad",
         "Romaine, Kalamata olives, feta cheese, red onion, cucumber, and tomatoes with lemon-oregano vinaigrette",
         12.99, True, "Vegetarian, Gluten-Free", 260),
        ("S004", "Salads", "Cobb Salad",
         "Chopped romaine with grilled chicken, avocado, hard-boiled egg, bacon, blue cheese, and tomatoes",
         16.99, True, "Gluten-Free", 520),
        ("S005", "Salads", "Strawberry Spinach Salad",
         "Baby spinach, fresh strawberries, candied walnuts, goat cheese, and balsamic vinaigrette",
         13.99, True, "Vegetarian", 310),

        # ---------- Main Course ----------
        ("M001", "Main Course", "Grilled Salmon",
         "Atlantic salmon fillet grilled with lemon-herb butter, served with seasonal vegetables and wild rice pilaf",
         26.99, True, "Seafood, Gluten-Free", 580),
        ("M002", "Main Course", "8oz Sirloin Steak",
         "USDA choice sirloin grilled to your preference, served with garlic mashed potatoes and asparagus",
         34.99, True, "Gluten-Free", 720),
        ("M003", "Main Course", "Chicken Parmesan",
         "Breaded chicken breast topped with housemade marinara and melted mozzarella, served with spaghetti",
         22.99, True, "None", 850),
        ("M004", "Main Course", "Shrimp Scampi",
         "Jumbo shrimp sauteed in garlic butter and white wine over linguine, finished with parsley and lemon",
         24.99, True, "Seafood", 640),
        ("M005", "Main Course", "Vegetable Stir-Fry",
         "Seasonal vegetables wok-tossed in ginger-soy sauce with tofu, served over steamed jasmine rice",
         17.99, True, "Vegetarian, Vegan", 420),
        ("M006", "Main Course", "BBQ Baby Back Ribs",
         "Full rack of slow-smoked pork ribs glazed with our signature BBQ sauce, served with coleslaw and cornbread",
         32.99, True, "Gluten-Free", 1100),
        ("M007", "Main Course", "Mushroom Risotto",
         "Arborio rice slow-cooked with porcini and cremini mushrooms, white wine, and Parmigiano-Reggiano",
         19.99, True, "Vegetarian, Gluten-Free", 560),
        ("M008", "Main Course", "Fish & Chips",
         "Beer-battered cod fillets fried golden, served with thick-cut fries, tartar sauce, and mushy peas",
         20.99, True, "Seafood", 780),

        # ---------- Sides ----------
        ("SI001", "Sides", "Garlic Mashed Potatoes",
         "Creamy russet potatoes whipped with roasted garlic, butter, and cream; topped with chives",
         5.99, True, "Vegetarian, Gluten-Free", 280),
        ("SI002", "Sides", "Truffle Parmesan Fries",
         "Crispy hand-cut fries tossed with truffle oil, grated Parmesan, and fresh herbs",
         7.99, True, "Vegetarian", 370),
        ("SI003", "Sides", "Steamed Seasonal Vegetables",
         "Market-fresh vegetables lightly steamed and finished with herb butter and sea salt",
         5.49, True, "Vegetarian, Vegan, Gluten-Free", 120),
        ("SI004", "Sides", "Mac & Cheese",
         "Cavatappi pasta in a rich three-cheese sauce topped with a toasted breadcrumb crust",
         6.99, True, "Vegetarian", 490),
        ("SI005", "Sides", "Cornbread",
         "Moist, lightly sweet cornbread baked in a cast-iron skillet, served warm with honey butter",
         4.49, True, "Vegetarian", 310),

        # ---------- Desserts ----------
        ("D001", "Desserts", "New York Cheesecake",
         "Dense, velvety cheesecake on a graham-cracker crust, topped with fresh berry compote",
         8.99, True, "Vegetarian", 490),
        ("D002", "Desserts", "Warm Chocolate Lava Cake",
         "Decadent dark chocolate cake with a molten center, served with vanilla bean ice cream and raspberry coulis",
         9.99, True, "Vegetarian", 580),
        ("D003", "Desserts", "Tiramisu",
         "Classic Italian dessert of espresso-soaked ladyfingers layered with mascarpone cream and dusted with cocoa",
         8.49, True, "Vegetarian", 430),
        ("D004", "Desserts", "Seasonal Fruit Sorbet",
         "Three scoops of house-made sorbet featuring seasonal fruit flavors; dairy-free and refreshing",
         6.99, True, "Vegan, Gluten-Free", 210),
        ("D005", "Desserts", "Creme Brulee",
         "Silky vanilla custard topped with a perfectly caramelized sugar crust; served with fresh berries",
         8.99, True, "Vegetarian, Gluten-Free", 390),

        # ---------- Beverages ----------
        ("B001", "Beverages", "Fresh Lemonade",
         "Hand-squeezed lemonade with cane sugar and fresh mint; still or sparkling",
         4.49, True, "Vegan, Gluten-Free", 130),
        ("B002", "Beverages", "Iced Tea",
         "House-brewed black tea served over ice with lemon; sweetened or unsweetened",
         3.49, True, "Vegan, Gluten-Free", 5),
        ("B003", "Beverages", "Mango Lassi",
         "Thick, chilled blend of Alphonso mango pulp, yogurt, honey, and a hint of cardamom",
         5.99, True, "Vegetarian, Gluten-Free", 240),
        ("B004", "Beverages", "Sparkling Water",
         "Chilled sparkling mineral water served with a slice of lemon or lime",
         2.99, True, "Vegan, Gluten-Free", 0),
        ("B005", "Beverages", "Craft Root Beer Float",
         "Premium craft root beer poured over two generous scoops of vanilla bean ice cream",
         6.49, True, "Vegetarian", 360),
    ]

    # Column headers
    headers = ["ID", "Category", "Name", "Description", "Price", "Available", "Dietary", "Calories"]

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menu"

    # Header styling
    header_font = Font(bold=True, size=12, color="000000")
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write and style headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write menu data
    for row_idx, item in enumerate(menu_items, start=2):
        for col_idx, value in enumerate(item, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 4))  # wrap Description

    # Auto-fit column widths
    column_widths = {
        1: 8,   # ID
        2: 14,  # Category
        3: 28,  # Name
        4: 55,  # Description
        5: 8,   # Price
        6: 10,  # Available
        7: 28,  # Dietary
        8: 10,  # Calories
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_path)
    print(f"Menu created successfully at: {output_path}  ({len(menu_items)} items)")


if __name__ == "__main__":
    create_menu_excel()
